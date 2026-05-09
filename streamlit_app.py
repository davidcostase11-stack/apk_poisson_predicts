import os
import requests
import pandas as pd
import numpy as np
import time
import re
import io
from io import StringIO
from scipy.stats import poisson
from datetime import datetime
import streamlit as st

# --- Page Config ---
st.set_page_config(page_title="Soccer Stats Predictor", layout="wide")

# --- Constants ---
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
# In the original script TODAY was fixed to 2026-05-04. Let's use that or allow user to pick.
DEFAULT_TODAY = datetime(2026, 5, 4)

LEAGUES_LIST = [
    "australia", "austria", "austria2", "belgium", "belgium2", "bosnia", "brazil", "bulgaria", "bulgaria2", "china",
    "croatia", "croatia2", "czechrepublic", "czechrepublic2", "denmark", "denmark2", "england", "england2", 
    "england3", "uae", "faroeislands", "finland", "finland2", "france", "france2", "germany", "germany2", "germany3", "georgia",
    "greece", "hungary", "hungary2", "iceland", "iceland2", "india", "indonesia", "ireland", "ireland2", "israel", "italy", "italy2", 
    "montenegro", "morocco", "netherlands", "netherlands2", "norway", "norway2", "poland", "poland2", "portugal", 
    "portugal2", "romania", "romania2", "singapore", "saudiarabia", "scotland", "scotland2", "slovakia", "slovenia", "southafrica", 
    "southkorea", "southkorea2", "spain", "spain2", "sweden", "sweden2", "switzerland", "thailand", "turkey", 
    "usa", "vietnam"
]

# --- Functions (from original script) ---

def get_with_retry(url, retries=3, backoff=2):
    for i in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=25)
            if r.status_code == 200: return r
        except: pass
        time.sleep(backoff * (i + 1))
    return None

def format_match_date(date_str):
    if not date_str or not isinstance(date_str, str): return date_str
    m_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
    try:
        m_pat = '|'.join(m_map.keys())
        match = re.search(fr'(\d{{1,2}})[/\s-]({m_pat}|\d{{1,2}})', date_str, re.IGNORECASE)
        if match:
            day = int(match.group(1))
            m_val = match.group(2).lower()
        else:
            match = re.search(fr'({m_pat})\s+(\d{{1,2}})', date_str, re.IGNORECASE)
            if match:
                m_val = match.group(1).lower()
                day = int(match.group(2))
            else: return date_str
        
        month = m_map[m_val] if m_val in m_map else int(m_val)
        year = 2026 if month < 7 else 2025
        return datetime(year, month, day)
    except: return date_str

def clean_team_name(name):
    if not isinstance(name, str): return str(name)
    name = re.sub(r'^\d+\.\s+', '', name)
    name = re.sub(r'\(.*?\)', '', name)
    return name.strip()

def extract_clean_results(df):
    df = df.dropna(axis=1, how='all')
    s_idx = -1
    for i in range(df.shape[1]):
        if df.iloc[:, i].astype(str).str.contains(r'\d\s?[:]\s?\d').any():
            s_idx = i
            break
    if s_idx == -1: return None

    d_idx = 0
    date_regex = r'\d{1,2}[/\s-](?:[a-zA-Z]{3}|\d{1,2})|(?:[a-zA-Z]{3})\s+(?:\d{1,2})'
    for i in range(min(s_idx, 4)):
        if df.iloc[:, i].astype(str).str.contains(date_regex).any():
            d_idx = i
            break

    rows = []
    for _, row in df.iterrows():
        res_val = str(row.iloc[s_idx])
        if re.search(r'\d\s?[:]\s?\d', res_val):
            rows.append({
                'Date_Raw': str(row.iloc[d_idx]),
                'Home Team': str(row.iloc[s_idx - 1]),
                'Result': res_val,
                'Away Team': str(row.iloc[s_idx + 1])
            })
    
    if not rows: return None
    
    res = pd.DataFrame(rows)
    res['Date_Obj'] = res['Date_Raw'].apply(format_match_date)
    res = res.sort_values(by='Date_Obj', na_position='first')
    res['Date'] = res['Date_Obj'].apply(lambda x: x.strftime('%d/%m/%Y') if isinstance(x, datetime) else "")
    res['Home Team'] = res['Home Team'].apply(clean_team_name)
    res['Away Team'] = res['Away Team'].apply(clean_team_name)
    return res

def extract_table_stats_by_index(df):
    try:
        data = df.iloc[1:].copy()
        stats = pd.DataFrame({
            'Team': data.iloc[:, 1].astype(str).str.strip(),
            'GP': pd.to_numeric(data.iloc[:, 2], errors='coerce'),
            'GF': pd.to_numeric(data.iloc[:, 6], errors='coerce'),
            'GA': pd.to_numeric(data.iloc[:, 7], errors='coerce')
        })
        return stats.dropna(subset=['Team', 'GP'])[stats['GP'] > 0]
    except: return None

def calculate_poisson_predictions(res_df, h_t, a_t):
    cols = ['Pred Score', 'Prob O1.5 Raw', 'Prob O1.5', 'Prob O2.5 Raw', 'Prob O2.5', 'Prob O3.5 Raw', 'Prob O3.5', 'Prob BTTS Raw', 'Prob BTTS', 'Pred Winner', 'Win Prob Raw', 'Win Prob', 'Prob 1X Raw', 'Prob 1X', 'Prob X2 Raw', 'Prob X2', 'Prob 12 Raw', 'Prob 12', 'Prob Home Raw', 'Prob Draw Raw', 'Prob Away Raw']
    if res_df is None or res_df.empty: return res_df
    
    if h_t is None or a_t is None or h_t.empty or a_t.empty:
        for c in cols: res_df[c] = None if 'Raw' not in c else 0.0
        return res_df
    
    avg_h_gf, avg_h_ga = h_t['GF'].sum() / h_t['GP'].sum(), h_t['GA'].sum() / h_t['GP'].sum()
    avg_a_gf, avg_a_ga = a_t['GF'].sum() / a_t['GP'].sum(), a_t['GA'].sum() / a_t['GP'].sum()
    
    h_t['Clean_Team'] = h_t['Team'].apply(clean_team_name)
    a_t['Clean_Team'] = a_t['Team'].apply(clean_team_name)
    h_s, a_s = h_t.set_index('Clean_Team'), a_t.set_index('Clean_Team')
    
    if avg_h_gf == 0: avg_h_gf = 0.01
    if avg_h_ga == 0: avg_h_ga = 0.01
    if avg_a_gf == 0: avg_a_gf = 0.01
    if avg_a_ga == 0: avg_a_ga = 0.01

    h_s['Atk'], h_s['Def'] = (h_s['GF'] / h_s['GP']) / avg_h_gf, (h_s['GA'] / h_s['GP']) / avg_h_ga
    a_s['Atk'], a_s['Def'] = (a_s['GF'] / a_s['GP']) / avg_a_gf, (a_s['GA'] / a_s['GP']) / avg_a_ga
    
    def pred(row):
        ht_raw, at_raw = str(row['Home Team']), str(row['Away Team'])
        ht, at = clean_team_name(ht_raw), clean_team_name(at_raw)
        h_atk = h_s.loc[ht, 'Atk'] if ht in h_s.index else 1.0
        h_def = h_s.loc[ht, 'Def'] if ht in h_s.index else 1.0
        a_atk = a_s.loc[at, 'Atk'] if at in a_s.index else 1.0
        a_def = a_s.loc[at, 'Def'] if at in a_s.index else 1.0
        l_h = h_atk * a_def * avg_h_gf
        l_a = a_atk * h_def * avg_a_gf
        l_h, l_a = max(0.01, l_h), max(0.01, l_a)
        m = np.outer(poisson.pmf(np.arange(8), l_h), poisson.pmf(np.arange(8), l_a))
        p_o15, p_o25, p_o35 = 1 - (m[0,0] + m[0,1] + m[1,0]), 1 - np.sum([m[i,j] for i in range(3) for j in range(3-i)]), 1 - np.sum([m[i,j] for i in range(4) for j in range(4-i)])
        p_btts = 1 - (np.sum(m[0, :]) + np.sum(m[:, 0]) - m[0,0])
        p_h, p_d, p_a = np.sum(np.tril(m, -1)), np.sum(np.diag(m)), np.sum(np.triu(m, 1))
        p_1x, p_x2, p_12 = p_h + p_d, p_a + p_d, p_h + p_a
        win_m = {'Home Win': p_h, 'Draw': p_d, 'Away Win': p_a}
        pw = max(win_m, key=win_m.get)
        hs, ascore = np.unravel_index(np.argmax(m), m.shape)
        return pd.Series([f"{hs}-{ascore}", p_o15, f"{p_o15:.1%}", p_o25, f"{p_o25:.1%}", p_o35, f"{p_o35:.1%}", p_btts, f"{p_btts:.1%}", pw, win_m[pw], f"{win_m[pw]:.1%}", p_1x, f"{p_1x:.1%}", p_x2, f"{p_x2:.1%}", p_12, f"{p_12:.1%}", p_h, p_d, p_a])
    
    res_df[cols] = res_df.apply(pred, axis=1)
    return res_df

def evaluate_accuracy(df):
    if df is None or 'Result' not in df.columns or 'Prob O1.5 Raw' not in df.columns: return [0]*8 + [0]
    def is_p(s):
        s = str(s).strip()
        if re.fullmatch(r'\d{2}:\d{2}', s) or any(x in s.lower() for x in ['pp', 'cancl', 'postp', 'sched']): return False
        return bool(re.search(r'\d+\s?[:\-]\s?\d+', s))
    def check(row, mt):
        res = str(row['Result'])
        if not is_p(res) or pd.isna(row.get('Prob O1.5 Raw')): return None
        match = re.search(r'(\d+)\s?[:\-]\s?(\d+)', res)
        if not match: return None
        h, a = int(match.group(1)), int(match.group(2))
        if mt == 'O1.5': return 1 if (h+a > 1.5) == (row['Prob O1.5 Raw'] > 0.5) else 0
        if mt == 'O2.5': return 1 if (h+a > 2.5) == (row['Prob O2.5 Raw'] > 0.5) else 0
        if mt == 'O3.5': return 1 if (h+a > 3.5) == (row['Prob O3.5 Raw'] > 0.5) else 0
        if mt == 'BTTS': return 1 if (h > 0 and a > 0) == (row['Prob BTTS Raw'] > 0.5) else 0
        if mt == 'Winner':
            act = 'Home Win' if h > a else ('Draw' if h == a else 'Away Win')
            return 1 if act == row['Pred Winner'] else 0
        if mt == '1X': return 1 if (h >= a) == (row['Prob 1X Raw'] > 0.5) else 0
        if mt == 'X2': return 1 if (a >= h) == (row['Prob X2 Raw'] > 0.5) else 0
        if mt == '12': return 1 if (h != a) == (row['Prob 12 Raw'] > 0.5) else 0
        return None
    metrics = ['O1.5', 'O2.5', 'O3.5', 'BTTS', 'Winner', '1X', 'X2', '12']
    for m in metrics: df[m+' Success'] = df.apply(lambda r: check(r, m), axis=1)
    played = df.dropna(subset=['Winner Success'])
    if len(played) > 0: return [played[m+' Success'].mean() * 100 for m in metrics] + [len(played)]
    return [0]*8 + [0]

def find_optimal_thresholds(df, raw_col, succ_col):
    if df is None or raw_col not in df.columns or succ_col not in df.columns: return None
    v = df.dropna(subset=[succ_col])
    if len(v) < 5: return None
    bl, bh = None, None
    for t in np.arange(0.01, 0.51, 0.01):
        sub = v[v[raw_col] <= t]
        if len(sub) >= 2:
            if sub[succ_col].mean() >= 0.95: bl = t
            else: break
    for t in np.arange(0.99, 0.49, -0.01):
        sub = v[v[raw_col] >= t]
        if len(sub) >= 2:
            if sub[succ_col].mean() >= 0.95: bh = t
            else: break
    return {'low': bl, 'high': bh}

def process_league(lg, writer, all_sum, all_strat, all_picks, target_date):
    # Results data
    r = get_with_retry(f"https://www.soccerstats.com/results.asp?league={lg}&pmtype=bydate")
    f_df = None
    if r:
        try:
            for table in pd.read_html(StringIO(r.text)):
                if len(table) > 10 and table.stack().astype(str).str.contains(r'\d\s?[:]\s?\d').any():
                    f_df = extract_clean_results(table); break
        except: pass
    
    if f_df is None: return False

    # Home/Away stats
    r = get_with_retry(f"https://www.soccerstats.com/homeaway.asp?league={lg}")
    ht, at = None, None
    if r:
        try:
            stats = []
            for table in pd.read_html(StringIO(r.text)):
                if len(table) > 5 and len(table.columns) >= 8 and 'gp' in [str(x).lower() for x in table.iloc[0].tolist()]: stats.append(table)
            if len(stats) >= 2: ht, at = extract_table_stats_by_index(stats[0]), extract_table_stats_by_index(stats[1])
        except: pass

    f_df = calculate_poisson_predictions(f_df, ht, at)
    accs = evaluate_accuracy(f_df)
    all_sum.append({'League': lg, 'Games': accs[8], 'O1.5': f"{accs[0]:.1f}%", 'O2.5': f"{accs[1]:.1f}%", 'O3.5': f"{accs[2]:.1f}%", 'BTTS': f"{accs[3]:.1f}%", 'Win': f"{accs[4]:.1f}%", '1X': f"{accs[5]:.1f}%", 'X2': f"{accs[6]:.1f}%", '12': f"{accs[7]:.1f}%"})
    
    strat = {'League': lg}
    mets_cfg = [('O1.5', 'Prob O1.5 Raw', 'Prob O1.5'), ('O2.5', 'Prob O2.5 Raw', 'Prob O2.5'), ('O3.5', 'Prob O3.5 Raw', 'Prob O3.5'), ('BTTS', 'Prob BTTS Raw', 'Prob BTTS'), ('Winner', 'Win Prob Raw', 'Win Prob'), ('1X', 'Prob 1X Raw', 'Prob 1X'), ('X2', 'Prob X2 Raw', 'Prob X2'), ('12', 'Prob 12 Raw', 'Prob 12')]
    l_t = {}
    for m_id, rc, d_n in mets_cfg:
        t = find_optimal_thresholds(f_df, rc, m_id+' Success')
        l_t[m_id] = t; res = []
        if t and t['low']: res.append(f"U: <{t['low']:.0%}")
        if t and t['high']: res.append(f"O: >{t['high']:.0%}")
        strat[m_id+' (95%)'] = " | ".join(res) if res else "No 95% zone"
    all_strat.append(strat)
    
    up = f_df[f_df['Date_Obj'] >= target_date]
    for _, m in up.iterrows():
        if pd.isna(m.get('Prob O1.5 Raw')): continue
        h_team, a_team = str(m['Home Team']).strip(), str(m['Away Team']).strip()
        h_gp = ht[ht['Team'] == h_team]['GP'].iloc[0] if ht is not None and not ht[ht['Team'] == h_team].empty else 0
        a_gp = at[at['Team'] == a_team]['GP'].iloc[0] if at is not None and not at[at['Team'] == a_team].empty else 0
        if h_gp < 5 or a_gp < 5: continue

        for m_id, rc, d_n in mets_cfg:
            t = l_t.get(m_id)
            if not t: continue
            v_p, trig, pt = m[rc], False, ""
            if t['low'] and v_p <= t['low']:
                if m_id.startswith('O'): trig, pt = True, f"UNDER {m_id}"
            if t['high'] and v_p >= t['high']:
                trig = True
                if m_id == 'Winner': pt = m['Pred Winner']
                elif m_id in ['1X', 'X2', '12']: pt = m_id
                else: pt = f"OVER {m_id}"
            if trig: all_picks.append({'Date': m['Date'], 'League': lg, 'Match': f"{m['Home Team']} vs {m['Away Team']}", 'Outcome': m_id, 'Prediction': pt, 'Prob': f"{v_p:.1%}", 'Safety': f">{t['high']:.0%}" if v_p >= (t['high'] or 1.1) else f"<{t['low']:.0%}"})
    
    s_name = lg[:31]
    save_df = f_df.drop(columns=[c for c in f_df.columns if 'Raw' in c or 'Success' in c or 'Date_Obj' in c])
    save_df.to_excel(writer, sheet_name=s_name, index=False)
    
    curr_row = len(save_df) + 3
    if ht is not None:
        pd.DataFrame([['HOME TABLE']]).to_excel(writer, sheet_name=s_name, startrow=curr_row, index=False, header=False)
        ht.to_excel(writer, sheet_name=s_name, startrow=curr_row+1, index=False)
        curr_row += len(ht) + 4
    if at is not None:
        pd.DataFrame([['AWAY TABLE']]).to_excel(writer, sheet_name=s_name, startrow=curr_row, index=False, header=False)
        at.to_excel(writer, sheet_name=s_name, startrow=curr_row+1, index=False)

    from openpyxl.styles import PatternFill
    ws, yellow = writer.sheets[s_name], PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    col_map = {name: i + 1 for i, name in enumerate(save_df.columns)}
    for r_idx, r_data in f_df.iterrows():
        excel_row = list(f_df.index).index(r_idx) + 2
        for m_id, rc, d_n in mets_cfg:
            t = l_t.get(m_id)
            if t and d_n in col_map:
                val = r_data[rc]
                if (t['low'] and val <= t['low']) or (t['high'] and val >= t['high']):
                    ws.cell(row=excel_row, column=col_map[d_n]).fill = yellow
    return True

# --- Streamlit UI ---

def main():
    st.title("⚽ Soccer Stats & Predictions")
    st.markdown("Scrape data from soccerstats.com and generate Poisson-based predictions.")

    with st.sidebar:
        st.header("Configuration")
        selected_leagues = st.multiselect("Select Leagues", LEAGUES_LIST, default=["england", "spain", "italy", "germany", "france", "portugal"])
        target_date = st.date_input("Filter matches from (inclusive)", DEFAULT_TODAY)
        target_date_dt = datetime.combine(target_date, datetime.min.time())
        
        run_button = st.button("Generate Predictions", type="primary")

    if run_button:
        if not selected_leagues:
            st.error("Please select at least one league.")
            return

        all_sum, all_strat, all_picks = [], [], []
        failed_leagues = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for i, lg in enumerate(selected_leagues):
                status_text.text(f"Processing {lg}... ({i+1}/{len(selected_leagues)})")
                success = process_league(lg, writer, all_sum, all_strat, all_picks, target_date_dt)
                if not success:
                    failed_leagues.append(lg)
                
                progress_bar.progress((i + 1) / len(selected_leagues))
                time.sleep(1.2) # To avoid getting blocked
            
            if failed_leagues:
                status_text.text(f"Retrying {len(failed_leagues)} failed leagues...")
                for lg in failed_leagues[:]:
                    success = process_league(lg, writer, all_sum, all_strat, all_picks, target_date_dt)
                    if success:
                        failed_leagues.remove(lg)
                    time.sleep(2)

            if all_sum: pd.DataFrame(all_sum).to_excel(writer, sheet_name='VALIDATION SUMMARY', index=False)
            if all_strat: pd.DataFrame(all_strat).to_excel(writer, sheet_name='BETTING STRATEGY', index=False)
            if all_picks:
                df_p = pd.DataFrame(all_picks)
                if not df_p.empty:
                    df_p['D_S'] = pd.to_datetime(df_p['Date'], format='%d/%m/%Y')
                    df_p.sort_values('D_S').drop(columns=['D_S']).to_excel(writer, sheet_name='TRIGGERED GAMES', index=False)

        status_text.success("Processing complete!")
        
        # Display Summaries
        if all_sum:
            st.subheader("Validation Summary")
            st.dataframe(pd.DataFrame(all_sum), use_container_width=True)
            
        if all_picks:
            st.subheader("Triggered Games (Predictions)")
            df_picks_show = pd.DataFrame(all_picks)
            if not df_picks_show.empty:
                st.dataframe(df_picks_show, use_container_width=True)
        
        if failed_leagues:
            st.warning(f"Could not process: {', '.join(failed_leagues)}")

        # Download Button
        st.download_button(
            label="Download Excel Results",
            data=output.getvalue(),
            file_name=f"soccer_predictions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    main()
